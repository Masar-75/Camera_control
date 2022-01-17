#!/usr/bin/python
import sys
sys.path.append('/home/pi/.local/lib/python3.7/site-packages')
import smbus
import math
import time
import openpyxl
from datetime import datetime
from datetime import timedelta
from threading import Thread

global str_folder 
global wb
global ws

str_folder = '/home/pi/Desktop/Data/'

# Power management registers
power_mgmt_1 = 0x6b
power_mgmt_2 = 0x6c
 
def read_byte(adr):
	return bus.read_byte_data(address, adr)
 
def read_word(adr):
	high = bus.read_byte_data(address, adr)
	low = bus.read_byte_data(address, adr+1)
	val = (high << 8) + low
	return val
 
def read_word_2c(adr):
	val = read_word(adr)
	if (val >= 0x8000):
		return -((65535 - val) + 1)
	else:
		return val

def magnitude(a,b,c):
    return math.sqrt((a*a)+(b*b)+(c*c))

def identify_gravity():
    duration=0
    res=0.030
    Nres=0
    t=0.0
    acc_x = 0.0
    acc_y = 0.0
    acc_z = 0.0
    while t <= duration:
        try:
            acc_x += read_word_2c(0x3b)
            acc_y += read_word_2c(0x3d)
            acc_z += read_word_2c(0x3f)
            Nres += 1
        except OSError:
            print("Exception OSERROR levee identify_gravity!")
        t += res
        time.sleep(res)
    if Nres == 0 :
        Nres = 1
    acc_x *= 9.81 / (Nres * 16384)
    acc_y *= 9.81 / (Nres * 16384)
    acc_z *= 9.81 / (Nres * 16384)
    return (acc_x, acc_y, acc_z)

def acc_magnitude_g():
    try:
        accel_xout = read_word_2c(0x3b)
        accel_yout = read_word_2c(0x3d)
        accel_zout = read_word_2c(0x3f)
    except OSError:
        accel_xout = 0
        accel_yout = 0
        accel_zout = 0
        print("Exception OSERROR levee acc_magnitude_g!")
    return magnitude(accel_xout, accel_yout, accel_zout) / 16384.0

def acc_magnitude_wo_gravity_ms2(xg_init, yg_init, zg_init):
    try:
        accel_xout = read_word_2c(0x3b)* 9.81 / 16384 - xg_init
        accel_yout = read_word_2c(0x3d)* 9.81 / 16384 - yg_init
        accel_zout = read_word_2c(0x3f)* 9.81 / 16384 - zg_init
    except OSError:
        accel_xout = 0
        accel_yout = 0
        accel_zout = 0
        print("Exception OSERROR levee dans acc_magnitude_wo_gravity_ms2!")	
    return magnitude(accel_xout, accel_yout, accel_zout)

while True:
    try:
        bus = smbus.SMBus(1) # or bus = smbus.SMBus(1) for Revision 2 boards
        address = 0x68 # This is the address value read via the i2cdetect command
         
        # Now wake the 6050 up as it starts in sleep mode
        bus.write_byte_data(address, power_mgmt_1, 0)
        break
    except OSError:
        continue

def get_max_accel(t_s):
    (xg, yg, zg) = identify_gravity()
    t=str((datetime.now()+timedelta(seconds=t_s)).strftime('%H:%M:%S'))
    max_accel = 0
    while t != str((datetime.now()+timedelta(seconds=0)).strftime('%H:%M:%S')):
        accel = acc_magnitude_wo_gravity_ms2(xg, yg, zg)
        if accel > max_accel:
            max_accel = accel
            #print("New max acceleration recorded for " + t + " : " + str(round(max_accel,3)) + "m/s2...")
        time.sleep(0.010)
    print("Last max acceleration recorded for " + t + " : " + str(round(max_accel,3)) + "m/s2!")
    return (t, max_accel)

def get_max_accel_minute():
    (xg, yg, zg) = identify_gravity()
    t=str(datetime.now().strftime('%H:%M'))
    max_accel = 0
    while t == str(datetime.now().strftime('%H:%M')):
        accel = acc_magnitude_wo_gravity_ms2(xg, yg, zg)
        if accel > max_accel:
            max_accel = accel
            #print("New max acceleration recorded for " + t + " : " + str(round(max_accel,3)) + "m/s2...")
        time.sleep(0.010)
    print("Last max acceleration recorded for " + t + " : " + str(round(max_accel,3)) + "m/s2!")
    return (t, max_accel)
    
def createTimeRange(wb, ws, inter_time_s):
    #inter_time_s = 4 #seconde - interval de mesure
    Nbr_interval_hr = int(60/inter_time_s*60)
    str_month = str(datetime.now().strftime('%B'))
    start_time_hour=20 #hr
    start_time_minute=00 #minutes
    Temps_enregistrement=16 #heures
    
    for i in range(Nbr_interval_hr*Temps_enregistrement):
        m = timedelta(seconds = i*inter_time_s)
        str_heure = str((datetime(2021,1,1,start_time_hour,start_time_minute)+m).strftime('%H:%M:%S'))
        ws.cell(row=2+i, column=1, value= str_heure)
        #print("cell row=", str(2+i), "; value = ", str_heure) 
    reference = 'Accel_data!$A$2:$A$'+str(1+Nbr_interval_hr*Temps_enregistrement)
    new_range = openpyxl.workbook.defined_name.DefinedName(str_month + "_TimeRange", attr_text=reference)
    wb.defined_names.append(new_range)
    return

def get_correct_col():
    if int(datetime.now().strftime('%H')) <= 12:
        dt = timedelta(days = 1)
    else:
        dt = timedelta(days = 0)
    nightdate = datetime.now() - dt
    str_Uniquedate = str(nightdate.strftime('%Y/%m/%d'))
    col=2
    while (ws.cell(row=1, column=col).value is not None) or (ws.cell(row=1, column=col).value is not None and ws.cell(row=1, column=col).value == str_Uniquedate):
        col+=1
    ws.cell(row=1, column = col, value= str_Uniquedate)
    return col

def time_left_before_start():
    hr_start = 20
    hr_end = 12
    hr = int(datetime.now().strftime('%H'))
    if hr > hr_start or hr < hr_end:
        _h = -1
    else:
        _h = int((20- int(datetime.now().strftime('%H'))-1)*60 + 60 - int(datetime.now().strftime('%M')))
        
    return _h
    
def are_we_into_recording_time(time_row):
    t=str(datetime.now().strftime('%H:%M:%S'))
    bool_now = False
    try:
        time_row[t]
        bool_now = True
    except:
        bool_now = False
    return bool_now

def get_time_row():
    time_row = {}
    str_month = str(datetime.now().strftime('%B'))
    rge_time = list(wb.defined_names[str_month + "_TimeRange"].destinations)[0][1].replace('$','')
    for cell in ws[rge_time]:
        time_row[cell[0].value]=cell[0].row
    return time_row

def wait_for_next_timerow(time_row):
    flag=1
    while flag:
        t=str(datetime.now().strftime('%H:%M:%S'))
        try:
            time_row[t]
            flag=0
        except:
            time.sleep(0.1)

def handle_new_year(str_last_year, wb):
    str_year = str(datetime.now().strftime('%Y'))
    if str_year != str_last_year:
        recFile = str_folder + 'Accel_data_beta_'+ str_year +'.xlsx'
        try:
            wb = openpyxl.load_workbook(filename = recFile)
            print("Workbook " + recFile + " opened with success!")
        except:
            wb = openpyxl.Workbook()
            print("New Workbook created!")
    return str_year, wb

def handle_new_month(str_last_month, wb, inter_time_s):
    str_month = str(datetime.now().strftime('%B'))  
    
    try:
        ws = wb[str_month + '_data']
    except:
        if str_month != str_last_month:
            ws = wb.create_sheet(str_month + '_data', 0)
            print("New month sheet '" + str_month + "_data' created!")
            createTimeRange(wb, ws, inter_time_s)
    
    return str_month, ws

def saving_file(obj, recfile):
    print("start saving process...")
    obj.save(recfile)
    print("File " + recFile + " saved!")
    
    
inter_time_s = 10 #secondes d'interval de mesure
init_time_row = True
str_month = "xxx"
str_year = "1970"
wb = openpyxl.Workbook()

while True:
    str_year, wb = handle_new_year(str_year, wb)
    str_month, ws = handle_new_month(str_month, wb, inter_time_s)
    
    if init_time_row:
        time_row = get_time_row()
        init_time_row = False
        
    if time_left_before_start() <= 0:
        col=get_correct_col()
        while True:
            wait_for_next_timerow(time_row)
            (h, accel) = get_max_accel(inter_time_s)
            try:
                Nrow = time_row[h]
            except:
                Nrow = 0
                print("timeline not found with h=" + h)
                break
            if Nrow:
                ws.cell(row=Nrow, column = col, value= accel)
                recFile = str_folder + 'Accel_data_beta_'+ str_year +'.xlsx'
                saving_thread=Thread(saving_file,args=(wb, recFile))
                saving_thread.start
                #print("File " + recFile + " saved!")
    else:
        time_left_min = time_left_before_start()
        print("sleep mode for " + str(time_left_min) + " minutes")
        time.sleep(time_left_min*60)







